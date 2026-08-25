"""Document ingestion.

An uploaded document is not "content we own". It is attacker-supplied text that
the model will later be asked to answer *from* — which makes ingestion a trust
boundary in its own right, with its own column in the severity matrix
(`ingest.document`) and its own posture.

Four things happen, in this order, and the order is the point:

    extract → ingest rails → chunk → index

Rails run on the whole document **before** it is chunked and written, so the
index never holds a raw identifier (`ingest.mask_before_index`, locked). A
document that fails a rail is **quarantined** rather than indexed with a flag
(`ingest.quarantine_on_block`, locked) — indexing it flagged would make
retrieval safety a matter of remembering to check the flag, every time, forever.

Retrieval is BM25 over chunks, gated by term coverage. Coverage is the honest
floor: BM25 ranks well but its scores are unbounded and corpus-relative, so
"is this chunk actually about the question" stays a separate, interpretable
number — the fraction of the query's terms the chunk contains.
"""

from __future__ import annotations

import json
import math
import re
import threading
import time
import uuid
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Iterable

TEXT_TYPES = {"txt", "md", "markdown", "csv", "json", "log", "text"}


class IngestError(ValueError):
    """Raised for anything the caller can fix: wrong type, empty file, too big."""


SHEET_TYPES = {"xlsx", "xlsm"}
IMAGE_TYPES = {"png", "jpg", "jpeg", "webp", "gif"}
BINARY_TYPES = {"pdf"} | SHEET_TYPES | IMAGE_TYPES

# What Claude accepts as an image block. `jpg` is spelled `jpeg` on the wire.
MEDIA_TYPE = {
    "png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
    "webp": "image/webp", "gif": "image/gif",
}

# A page with a text layer yields far more than this. Below it, the page is a
# picture of text — a scan — and needs transcribing instead of parsing.
TEXT_LAYER_MIN_CHARS = 40
OCR_TARGET_WIDTH = 1500          # px; ~150 dpi on A4, well inside the image limit


class OCRUnavailable(IngestError):
    """A scan arrived and nothing can read it. Says so rather than indexing silence."""


@dataclass
class Extraction:
    """Text, and an honest account of where it came from."""

    text: str
    method: str          # paste · text · pdf.text · pdf.ocr · pdf.mixed · sheet · image.ocr
    pages: int = 0
    transcribed: int = 0

    @property
    def detail(self) -> str:
        if self.method == "pdf.text":
            return f"text layer, {self.pages} pages"
        if self.method in ("pdf.ocr", "pdf.mixed"):
            return f"{self.transcribed} of {self.pages} pages transcribed"
        if self.method == "image.ocr":
            return "image transcribed"
        if self.method == "sheet":
            return f"{self.pages} sheets"
        return self.method


def extension(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else "txt"


def _decode(data: bytes) -> str:
    for encoding in ("utf-8", "utf-8-sig", "cp1252"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _rows_as_markdown_table(rows: list[list[str]]) -> str:
    """Already-normalised string rows as one markdown pipe table, first row as
    header. Shared by the spreadsheet and PDF-table extraction paths — same
    shape, same reason: the same text has to serve BM25 retrieval *and* be
    readable when the model quotes it back, and a table only reads correctly
    if the columns still line up."""
    rows = [r for r in rows if any(r)]
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    head, body = rows[0], rows[1:]
    lines = ["| " + " | ".join(head) + " |",
             "|" + "|".join(["---"] * width) + "|"]
    lines += ["| " + " | ".join(r) + " |" for r in body]
    return "\n".join(lines)


def _sheet_text(data: bytes, max_rows: int = 5000) -> tuple[str, int]:
    """A workbook as markdown tables, one per sheet.

    Tables rather than a flat dump: the same rows have to serve BM25 retrieval
    *and* be readable when the model quotes them back.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise IngestError("Spreadsheet ingestion needs openpyxl — `pip install openpyxl`.") from exc

    import io

    book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    rows_seen = 0
    for sheet in book.worksheets:
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            if rows_seen >= max_rows:
                parts.append("[truncated — sheet exceeds the row limit]")
                break
            cells = ["" if c is None else str(c).strip().replace("|", "\\|") for c in row]
            if not any(cells):
                continue
            rows.append(cells)
            rows_seen += 1
        table = _rows_as_markdown_table(rows)
        if table:
            parts.append(f"## {sheet.title}\n\n" + table)
    book.close()
    return "\n\n".join(parts), len(book.worksheets)


def _pdf_pages_text(data: bytes) -> list[str]:
    """One string per page — tables rendered as markdown pipe tables wherever
    PyMuPDF can find one, everything else as plain paragraph text, merged
    back in the order it actually appears on the page.

    A flat `extract_text()` call has no notion of a table at all: a table's
    cells come back in whatever order the PDF's content stream happens to
    list them, which routinely isn't reading order — a two-column table can
    interleave into "Fee ScheduleRupees" instead of a row a person or a
    retrieval index can use. Detecting each table's own region and rendering
    it separately is what keeps the columns lined up.
    """
    import fitz  # PyMuPDF — already a dependency, for the scan-rasterising path below

    doc = fitz.open(stream=data, filetype="pdf")
    try:
        return [_pdf_page_text(page) for page in doc]
    finally:
        doc.close()


def _pdf_page_text(page: Any) -> str:
    import fitz

    try:
        tables = list(page.find_tables())
    except Exception:  # noqa: BLE001 — a table-detection failure must not sink the page
        tables = []
    if not tables:
        return page.get_text("text").strip()

    # Everything on the page that is not inside a detected table's own
    # region, plus each table rendered as markdown — both timestamped by
    # vertical position, so the merge reads in the order the page does
    # rather than "all the prose, then all the tables".
    table_rects = [fitz.Rect(t.bbox) for t in tables]
    items: list[tuple[float, str]] = []
    for block in page.get_text("blocks"):
        rect = fitz.Rect(block[:4])
        if any(rect.intersects(tr) for tr in table_rects):
            continue  # this text is the table below, already captured there
        text = str(block[4]).strip()
        if text:
            items.append((rect.y0, text))
    for table, rect in zip(tables, table_rects):
        rows = [["" if c is None else str(c).strip().replace("|", "\\|") for c in row]
               for row in table.extract()]
        md = _rows_as_markdown_table(rows)
        if md:
            items.append((rect.y0, md))
    items.sort(key=lambda item: item[0])
    return "\n\n".join(text for _, text in items).strip()


def _rasterise(data: bytes, index: int) -> tuple[bytes, str]:
    """One PDF page as a PNG, sized for the vision call."""
    try:
        import fitz  # PyMuPDF
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise OCRUnavailable(
            "Reading a scanned PDF needs PyMuPDF — `pip install pymupdf`."
        ) from exc

    doc = fitz.open(stream=data, filetype="pdf")
    try:
        page = doc[index]
        # Scale so the long edge lands near OCR_TARGET_WIDTH: bigger costs more
        # and reads no better; smaller starts losing small print.
        zoom = min(3.0, max(1.0, OCR_TARGET_WIDTH / max(1.0, page.rect.width)))
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
        return pix.tobytes("png"), "image/png"
    finally:
        doc.close()


def extract(filename: str, data: bytes, *, allowed: Iterable[str] = (),
            ocr: Callable[[bytes, str, str], str] | None = None,
            max_ocr_pages: int = 20) -> Extraction:
    """Bytes → text, by whichever route the file actually needs.

    `ocr` is injected rather than imported so this module never depends on a
    model client. When it is missing, a scan fails loudly: an empty document
    that indexed cleanly would be worse than a refusal.
    """
    ext = extension(filename)
    allowed = {str(a).lower() for a in allowed} or TEXT_TYPES | BINARY_TYPES
    if ext not in allowed:
        raise IngestError(
            f".{ext} is not an accepted type. Accepted: {', '.join(sorted(allowed))}"
        )

    # ---- spreadsheets ------------------------------------------------
    if ext in SHEET_TYPES:
        text, sheets = _sheet_text(data)
        if not text.strip():
            raise IngestError("That workbook has no readable cells.")
        return Extraction(text, "sheet", pages=sheets)

    # ---- images ------------------------------------------------------
    if ext in IMAGE_TYPES:
        if ocr is None:
            raise OCRUnavailable(
                "Reading an image needs a model. Set ANTHROPIC_API_KEY and restart, "
                "or paste the text instead."
            )
        text = ocr(data, MEDIA_TYPE[ext], "Transcribe this image.")
        if not text.strip():
            raise IngestError("No readable text in that image.")
        return Extraction(text, "image.ocr", pages=1, transcribed=1)

    # ---- pdf: text layer first, transcription only where it is missing --
    if ext == "pdf":
        pages = _pdf_pages_text(data)
        thin = [i for i, t in enumerate(pages) if len(t) < TEXT_LAYER_MIN_CHARS]
        if not thin:
            return Extraction("\n\n".join(pages), "pdf.text", pages=len(pages))

        if ocr is None:
            if any(pages):
                # Some pages read; say what was lost rather than pretending.
                return Extraction("\n\n".join(p for p in pages if p), "pdf.text",
                                  pages=len(pages))
            raise OCRUnavailable(
                f"That PDF is a scan — {len(pages)} pages with no text layer. Reading it "
                "needs a model: set ANTHROPIC_API_KEY and restart."
            )

        budget = list(thin)[:max_ocr_pages]
        done = 0
        for i in budget:
            image, media = _rasterise(data, i)
            pages[i] = ocr(image, media, f"Transcribe page {i + 1}.")
            done += 1
        if len(thin) > len(budget):
            pages.append(f"[{len(thin) - len(budget)} further scanned pages were not "
                         "transcribed — ingest.ocr_max_pages]")
        method = "pdf.ocr" if len(thin) == len(pages) else "pdf.mixed"
        text = "\n\n".join(p for p in pages if p.strip())
        if not text.strip():
            raise IngestError("Nothing readable came out of that PDF.")
        return Extraction(text, method, pages=len(pages), transcribed=done)

    # ---- plain text --------------------------------------------------
    return Extraction(_decode(data), "text")


# ---------------------------------------------------------------------------
# Chunking
# ---------------------------------------------------------------------------
_PARA = re.compile(r"\n\s*\n")
_SENTENCE = re.compile(r"(?<=[.!?])\s+")


def chunk_text(text: str, size: int = 700, overlap: int = 80) -> list[str]:
    """Paragraph-packed chunks of roughly `size` characters, with overlap.

    Paragraphs are the natural unit, so they are packed whole where they fit. A
    paragraph longer than `size` is split on sentence boundaries; a sentence
    longer than `size` is cut on width, because something has to give and a
    silent drop is worse than a hard cut.
    """
    text = text.strip()
    if not text:
        return []
    size = max(120, int(size))
    overlap = max(0, min(int(overlap), size // 2))

    units: list[str] = []
    for para in _PARA.split(text):
        para = " ".join(para.split())
        if not para:
            continue
        if len(para) <= size:
            units.append(para)
            continue
        buf = ""
        for sentence in _SENTENCE.split(para):
            while len(sentence) > size:
                units.append(sentence[:size])
                sentence = sentence[size:]
            if not buf:
                buf = sentence
            elif len(buf) + 1 + len(sentence) <= size:
                buf = f"{buf} {sentence}"
            else:
                units.append(buf)
                buf = sentence
        if buf:
            units.append(buf)

    chunks: list[str] = []
    current = ""
    for unit in units:
        if not current:
            current = unit
        elif len(current) + 2 + len(unit) <= size:
            current = f"{current}\n\n{unit}"
        else:
            chunks.append(current)
            tail = current[-overlap:] if overlap else ""
            # Start the overlap at a word boundary so a chunk never opens mid-word.
            if tail and " " in tail:
                tail = tail[tail.index(" ") + 1:]
            current = f"{tail}\n\n{unit}".strip() if tail else unit
    if current:
        chunks.append(current)
    return chunks


# ---------------------------------------------------------------------------
# Tokenising and BM25
# ---------------------------------------------------------------------------
_STOP = {
    "the", "a", "an", "is", "are", "was", "were", "to", "of", "and", "or", "in",
    "on", "for", "with", "my", "i", "do", "does", "how", "what", "can", "need",
    "you", "your", "me", "it", "this", "that", "be", "have", "has", "at", "from",
    # Function words that carry no topic but inflate the coverage denominator.
    # "Where do I file a grievance and how soon will someone respond?" scored
    # 0.143 against a 0.15 gate purely on these, and the right document was
    # dropped.
    "where", "when", "who", "why", "which", "will", "would", "should", "could",
    "shall", "may", "must", "about", "any", "some", "someone", "somebody",
    "anyone", "there", "here", "get", "got", "much", "many", "long", "soon",
    "please", "tell", "give", "want", "know", "into", "out", "if", "so", "than",
    "then", "them", "they", "their", "our", "its", "his", "her", "been", "being",
    "am", "not", "but", "also", "just", "like", "make", "made", "take", "does",
}
_TOKEN = re.compile(r"[a-z0-9]+")

# Light suffix stripping, applied to queries and documents alike. Not a full
# Porter stemmer — just enough that "file", "filed" and "filing" are one term.
# Without it, coverage measures vocabulary rather than topic.
_SUFFIXES = ("ations", "ation", "ements", "ement", "ments", "ment", "ances",
             "ance", "ings", "ing", "ies", "ied", "als", "al", "es", "ed", "s")


def stem(word: str) -> str:
    if len(word) <= 3 or word.isdigit():
        return word
    for suffix in _SUFFIXES:
        if word.endswith(suffix) and len(word) - len(suffix) >= 3:
            base = word[: -len(suffix)]
            if suffix in ("ies", "ied"):
                base += "y"
            word = base
            break
    # A trailing "e" is the last difference between file / filing / filed.
    return word[:-1] if len(word) > 3 and word.endswith("e") else word


#: A question longer than this must match at least two distinct terms, not
#: just clear the coverage ratio. Below it, one match is all a question has
#: to give — "renewal fee?" is two terms and entirely legitimate.
MIN_TERMS_FOR_PAIR = 3

K1 = 1.5
B = 0.75


def tokens(text: str) -> list[str]:
    return [stem(t) for t in _TOKEN.findall(text.lower())
            if t not in _STOP and len(t) > 2]


# ---------------------------------------------------------------------------
# Documents
# ---------------------------------------------------------------------------
@dataclass
class Document:
    """One ingested document and the record of what ingesting it found."""

    id: str
    title: str
    source: str = "upload"          # filename, "paste", or "built-in"
    kind: str = "txt"
    added_at: float = field(default_factory=time.time)
    chars: int = 0
    chunks: list[str] = field(default_factory=list)
    status: str = "indexed"         # indexed | quarantined
    verdict: str = "pass"
    masked: int = 0
    findings: list[dict[str, Any]] = field(default_factory=list)
    reason: str = ""
    request_id: str = ""
    method: str = "text"        # how the text was obtained; see Extraction
    #: True only when this entry is the output of a real `Engine.ingest()` call
    #: — rails actually ran, `verdict` and `masked` are real numbers, not
    #: defaults. False for a built-in seeded straight into the store before an
    #: `Engine` (and its rails) exist; `Engine.__init__` finds those and
    #: re-ingests them for real. A fast, dependency-free `Corpus(seed=True)`
    #: with no `Engine` at all — plenty of tests want exactly that — never
    #: flips this, and that is the honest answer for it: nothing rail-checked
    #: this content, so nothing should claim to have.
    rails_applied: bool = False

    @property
    def indexed(self) -> bool:
        return self.status == "indexed"

    def to_dict(self, *, with_chunks: bool = False) -> dict[str, Any]:
        d = {
            "id": self.id,
            "title": self.title,
            "source": self.source,
            "kind": self.kind,
            "added_at": self.added_at,
            "chars": self.chars,
            "n_chunks": len(self.chunks),
            "status": self.status,
            "verdict": self.verdict,
            "masked": self.masked,
            "findings": self.findings,
            "reason": self.reason,
            "request_id": self.request_id,
            "method": self.method,
            "built_in": self.source == "built-in",
            "rails_applied": self.rails_applied,
        }
        if with_chunks:
            d["chunks"] = self.chunks
        return d

    @classmethod
    def from_dict(cls, d: dict[str, Any]) -> Document:
        return cls(
            id=d["id"], title=d["title"], source=d.get("source", "upload"),
            kind=d.get("kind", "txt"), added_at=d.get("added_at", time.time()),
            chars=d.get("chars", 0), chunks=list(d.get("chunks") or []),
            status=d.get("status", "indexed"), verdict=d.get("verdict", "pass"),
            masked=d.get("masked", 0), findings=list(d.get("findings") or []),
            reason=d.get("reason", ""), request_id=d.get("request_id", ""),
            method=d.get("method", "text"),
            rails_applied=bool(d.get("rails_applied", False)),
        )


@dataclass
class Hit:
    doc_id: str
    title: str
    chunk_index: int
    text: str
    score: float          # bm25, corpus-relative
    coverage: float       # fraction of query terms present — the interpretable one

    def as_context(self) -> str:
        return f"{self.title}: {self.text}"


# ---------------------------------------------------------------------------
# Corpus
# ---------------------------------------------------------------------------
class Corpus:
    """The document store and its index.

    Small enough to keep in memory and on one JSON file. The interface — `add`,
    `remove`, `search` — is the same shape a vector store would offer, so
    swapping the ranking out does not touch the engine.
    """

    def __init__(self, path: str | Path | None = None, seed: bool = True) -> None:
        self.path = Path(path) if path else None
        self._docs: dict[str, Document] = {}
        self._lock = threading.RLock()
        # index state, rebuilt on every mutation
        self._postings: list[tuple[str, int, Counter, int]] = []
        self._df: Counter = Counter()
        self._avgdl: float = 1.0
        # Which built-ins this store has ever been given. Not the same as which
        # it currently holds: an operator may have deleted one on purpose.
        self._seeds_installed: set[str] = set()
        if self.path and self.path.exists():
            self.load()
        elif seed:
            self.seed_builtin()

    # ---- persistence -------------------------------------------------
    def load(self) -> None:
        assert self.path is not None
        try:
            raw = json.loads(self.path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            # A corrupt store must not take the server down — start clean and
            # let the next save overwrite it.
            self.seed_builtin()
            return
        with self._lock:
            self._docs = {
                d["id"]: Document.from_dict(d) for d in raw.get("documents", [])
            }
            installed = raw.get("seeds_installed")
            # A store written before this field existed has, by definition, been
            # given every built-in that existed when it was written — which is
            # exactly the set of seed documents it still holds.
            self._seeds_installed = set(installed) if installed is not None else {
                i for i in self._docs if i.startswith("seed:")
            }
            self._reindex()
        if not self._docs:
            self.seed_builtin()
        else:
            self.install_new_builtins()

    def save(self) -> None:
        if self.path is None:
            return
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with self._lock:
            payload = {
                "version": 1,
                "updated_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
                "documents": [d.to_dict(with_chunks=True) for d in self._docs.values()],
                "seeds_installed": sorted(self._seeds_installed),
            }
        tmp = self.path.with_suffix(".tmp")
        tmp.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        tmp.replace(self.path)

    def install_new_builtins(self) -> int:
        """Add built-in documents this store has never been given.

        A shipped corpus update has to reach deployments that already exist,
        and the only alternative — wiping the store — would take the operator's
        uploads with it. Tracked by id rather than by presence, so a built-in
        somebody deleted on purpose stays deleted instead of returning at every
        restart.
        """
        from .seed import CORPUS

        added = 0
        with self._lock:
            for doc in CORPUS:
                doc_id = f"seed:{doc['id']}"
                if doc_id in self._seeds_installed:
                    continue
                self._seeds_installed.add(doc_id)
                if doc_id in self._docs:
                    continue
                self._docs[doc_id] = Document(
                    id=doc_id, title=doc["title"], source="built-in", kind="txt",
                    chars=len(doc["text"]), chunks=[doc["text"]],
                    status="indexed", verdict="pass",
                )
                added += 1
            if added:
                self._reindex()
        if added:
            self.save()
        return added

    def seed_builtin(self) -> None:
        """Install whatever `CORPUS` currently holds — nothing, since the
        built-in demo documents were removed. Kept rather than deleted: a
        deployment that wants a starter knowledge base again populates
        `CORPUS` and this path installs it the same way it always did.
        """
        from .seed import CORPUS

        with self._lock:
            for doc in CORPUS:
                d = Document(
                    id=f"seed:{doc['id']}", title=doc["title"], source="built-in",
                    kind="txt", chars=len(doc["text"]), chunks=[doc["text"]],
                    status="indexed", verdict="pass",
                )
                self._docs.setdefault(d.id, d)
                self._seeds_installed.add(d.id)
            self._reindex()

    # ---- mutation ----------------------------------------------------
    def add(self, doc: Document) -> Document:
        with self._lock:
            self._docs[doc.id] = doc
            self._reindex()
        self.save()
        return doc

    def remove(self, doc_id: str) -> bool:
        with self._lock:
            gone = self._docs.pop(doc_id, None) is not None
            if gone:
                self._reindex()
        if gone:
            self.save()
        return gone

    def reset(self) -> None:
        """Back to the built-in set. Used by tests and the console's reset."""
        with self._lock:
            self._docs.clear()
            self.seed_builtin()
        self.save()

    # ---- reading -----------------------------------------------------
    def all(self) -> list[Document]:
        with self._lock:
            return sorted(self._docs.values(), key=lambda d: (d.source != "built-in", -d.added_at))

    def get(self, doc_id: str) -> Document | None:
        return self._docs.get(doc_id)

    def stats(self) -> dict[str, Any]:
        docs = self.all()
        indexed = [d for d in docs if d.indexed]
        return {
            "documents": len(docs),
            "indexed": len(indexed),
            "quarantined": sum(1 for d in docs if d.status == "quarantined"),
            "uploaded": sum(1 for d in docs if d.source != "built-in"),
            "chunks": sum(len(d.chunks) for d in indexed),
            "masked_values": sum(d.masked for d in docs),
        }

    # ---- index -------------------------------------------------------
    def _reindex(self) -> None:
        postings: list[tuple[str, int, Counter, int]] = []
        df: Counter = Counter()
        total = 0
        for doc in self._docs.values():
            if not doc.indexed:
                continue  # quarantined documents are not searchable. That is the point.
            for i, chunk in enumerate(doc.chunks):
                tf = Counter(tokens(f"{doc.title} {chunk}"))
                if not tf:
                    continue
                length = sum(tf.values())
                postings.append((doc.id, i, tf, length))
                total += length
                for term in tf:
                    df[term] += 1
        self._postings = postings
        self._df = df
        self._avgdl = (total / len(postings)) if postings else 1.0

    def search(self, query: str, k: int = 4, min_coverage: float = 0.15) -> list[Hit]:
        """BM25 ranking, term-coverage gate, and a floor of two matched terms.

        Coverage is checked before ranking because the two answer different
        questions: BM25 says "which of these is most relevant", coverage says
        "is any of this about the question at all". A corpus-relative score
        cannot answer the second one.
        """
        q = tokens(query)
        if not q:
            return []
        q_set = set(q)
        with self._lock:
            postings, df, avgdl = self._postings, self._df, self._avgdl
            n = len(postings)
            if not n:
                return []
            hits: list[Hit] = []
            for doc_id, idx, tf, length in postings:
                matched = q_set & set(tf)
                coverage = len(matched) / len(q_set)
                if coverage < min_coverage:
                    continue
                # Coverage is a ratio, so a short question needs proportionally
                # fewer matches to clear the gate: five terms need one. That let
                # "how do I apply for a fishing permit" reach a trade-licence
                # chunk on the word "applying" alone. One shared term is a
                # coincidence, not a topic — unless the question is that short.
                if len(matched) < 2 and len(q_set) > MIN_TERMS_FOR_PAIR:
                    continue
                score = 0.0
                for term in q_set:
                    f = tf.get(term, 0)
                    if not f:
                        continue
                    idf = math.log(1 + (n - df[term] + 0.5) / (df[term] + 0.5))
                    score += idf * (f * (K1 + 1)) / (f + K1 * (1 - B + B * length / avgdl))
                doc = self._docs[doc_id]
                hits.append(Hit(doc_id, doc.title, idx, doc.chunks[idx], score, coverage))
        hits.sort(key=lambda h: (-h.score, -h.coverage))
        return hits[:k]


def new_document_id(title: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")[:40] or "document"
    return f"{slug}-{uuid.uuid4().hex[:6]}"
